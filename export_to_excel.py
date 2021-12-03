"""This module provides the management of exporting to xlsx files"""

from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import database
import os
import tempfile


def open_excel(location, from_date, to_date):
    # Get the users temporary folder location
    temp_file = tempfile.gettempdir()

    # Query the database for the monthly data
    data_list = database.fetch_history_records(location, from_date, to_date)

    # Open a new Excel instance and activate the first sheet
    wb = Workbook()
    sht = wb.active

    # This is the row to begin putting data on in Excel
    i_row = 1

    # Add data to the excel document
    try:
        # Add headers to the Excel spreadsheet
        sht.cell(row=i_row, column=1).value = 'Weather Location'
        sht.cell(row=i_row, column=2).value = 'Record Date'
        sht.cell(row=i_row, column=3).value = 'Temp High (F)'
        sht.cell(row=i_row, column=4).value = 'Temp Average (F)'
        sht.cell(row=i_row, column=5).value = 'Temp Low (F)'
        sht.cell(row=i_row, column=6).value = 'Dew Point High (F)'
        sht.cell(row=i_row, column=7).value = 'Dew Point Average (F)'
        sht.cell(row=i_row, column=8).value = 'Dew Point Low (F)'
        sht.cell(row=i_row, column=9).value = 'Humidity High (%)'
        sht.cell(row=i_row, column=10).value = 'Humidity Average (%)'
        sht.cell(row=i_row, column=11).value = 'Humidity Low (%)'
        sht.cell(row=i_row, column=12).value = 'Speed High (mph)'
        sht.cell(row=i_row, column=13).value = 'Speed Average (mph)'
        sht.cell(row=i_row, column=14).value = 'Speed Low (mph)'
        sht.cell(row=i_row, column=15).value = 'Pressure High (in)'
        sht.cell(row=i_row, column=16).value = 'Pressure Low (in)'
        sht.cell(row=i_row, column=17).value = 'Precipitation (in)'

        # Adjust column size based on length of header value
        for x in range(1, 18):
            column_title = sht.cell(row=1, column=x).value
            sht.column_dimensions[get_column_letter(x)].width = len(column_title)

        # Populate daily values
        i_row += 1
        for x in data_list:
            sht.cell(row=i_row, column=1).value = x[1]  # Website alias
            sht.cell(row=i_row, column=2).value = x[2]  # Record Date
            sht.cell(row=i_row, column=3).value = x[3]  # Temp High
            sht.cell(row=i_row, column=4).value = x[4]  # Temp Average
            sht.cell(row=i_row, column=5).value = x[5]  # Temp Low
            sht.cell(row=i_row, column=6).value = x[6]  # Dew Point High
            sht.cell(row=i_row, column=7).value = x[7]  # Dew Point Average
            sht.cell(row=i_row, column=8).value = x[8]  # Dew Point Low
            sht.cell(row=i_row, column=9).value = x[9]  # Humidity High
            sht.cell(row=i_row, column=10).value = x[10]  # Humidity Average
            sht.cell(row=i_row, column=11).value = x[11]  # Humidity Low
            sht.cell(row=i_row, column=12).value = x[12]  # Speed High
            sht.cell(row=i_row, column=13).value = x[13]  # Speed Average
            sht.cell(row=i_row, column=14).value = x[14]  # Speed Low
            sht.cell(row=i_row, column=15).value = x[15]  # Pressure High
            sht.cell(row=i_row, column=16).value = x[16]  # Pressure Low
            sht.cell(row=i_row, column=17).value = x[17]  # Precipitation

            i_row += 1

    except Exception as e:
        print("An unexpected error occurred")
        print(type(e), e)

    # Save the Excel file to the temporary folder
    excel_file = os.path.join(temp_file, 'Precipitation_Data.xlsx')
    wb.save(excel_file)

    # Open the file for the user to see the results
    # Note: This will use whatever the default application for .xlsx files is (ie. Excel, LibreOffice, etc.)
    os.startfile(excel_file)

