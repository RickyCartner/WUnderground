"""This module provides the management of exporting to xlsx files"""

from openpyxl.utils import get_column_letter
from openpyxl import Workbook

import os
import tempfile


def open_excel(table_data):
    # Get the users temporary folder location
    temp_file = tempfile.gettempdir()

    # Query the database for the monthly data
    # data_list = database.fetch_history_records(location, from_date, to_date)

    # Open a new Excel instance and activate the first sheet
    wb = Workbook()
    sht = wb.active

    # This is the row to begin putting data on in Excel
    i_row = 1
    # self.main_ui.txtHeatIndexHigh.setText(str(table_data.get('Heat Index High')))
    # Add data to the excel document
    try:
        # Add headers to the Excel spreadsheet
        sht.cell(row=i_row, column=1).value = 'Weather Location'
        sht.cell(row=i_row, column=2).value = 'Record Date'
        sht.cell(row=i_row, column=3).value = 'Temp High (F)'
        sht.cell(row=i_row, column=4).value = 'Temp Low (F)'
        sht.cell(row=i_row, column=5).value = 'Temp Average (F)'
        sht.cell(row=i_row, column=6).value = 'Dew Point High (F)'
        sht.cell(row=i_row, column=7).value = 'Dew Point Low (F)'
        sht.cell(row=i_row, column=8).value = 'Dew Point Average (F)'
        sht.cell(row=i_row, column=9).value = 'Heat Index High (%)'
        sht.cell(row=i_row, column=10).value = 'Heat Index Low (%)'
        sht.cell(row=i_row, column=11).value = 'Heat Index Average (%)'
        sht.cell(row=i_row, column=12).value = 'Speed High (mph)'
        sht.cell(row=i_row, column=13).value = 'Speed Low (mph)'
        sht.cell(row=i_row, column=14).value = 'Speed Average (mph)'
        sht.cell(row=i_row, column=15).value = 'Gust High (mph)'
        sht.cell(row=i_row, column=16).value = 'Gust Low (mph)'
        sht.cell(row=i_row, column=17).value = 'Gust Average (mph)'
        sht.cell(row=i_row, column=18).value = 'Wind Chill High (mph)'
        sht.cell(row=i_row, column=19).value = 'Wind Chill Low (mph)'
        sht.cell(row=i_row, column=20).value = 'Wind Chill Average (mph)'
        sht.cell(row=i_row, column=21).value = 'Pressure Min (in)'
        sht.cell(row=i_row, column=22).value = 'Pressure Max (in)'
        sht.cell(row=i_row, column=23).value = 'Pressure Trend (in)'
        sht.cell(row=i_row, column=24).value = 'Precipitation Rate (in)'
        sht.cell(row=i_row, column=25).value = 'Precipitation Total (in)'

        # Adjust column size based on length of header value
        for x in range(1, 25):
            column_title = sht.cell(row=1, column=x).value
            sht.column_dimensions[get_column_letter(x)].width = len(column_title)

        # Populate daily values
        for row, value in enumerate(table_data):
            for col, (k, v) in enumerate(value.items()):
                sht.cell(row=row+2, column=col+1).value = v  # Website alias

    except Exception as e:
        print("An unexpected error occurred")
        print(type(e), e)

    # Save the Excel file to the temporary folder
    excel_file = os.path.join(temp_file, 'Precipitation_Data.xlsx')
    wb.save(excel_file)

    # Open the file for the user to see the results
    # Note: This will use whatever the default application for .xlsx files is (ie. Excel, LibreOffice, etc.)
    os.startfile(excel_file)

